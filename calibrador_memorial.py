# -*- coding: utf-8 -*-
"""
CALIBRADOR DO MEMORIAL — Morais Engenharia
Ferramenta de ajuste visual para assinatura e checkbox do Memorial Excel.

USO:
  1. Execute: python calibrador_memorial.py
  2. Selecione o Memorial (.xls ou .xlsx) e a imagem de assinatura
  3. Ajuste os sliders/campos até o posicionamento ficar correto
  4. Clique "GERAR PREVIEW" para ver o resultado no Excel
  5. Copie os valores finais para o app.py principal

DEPENDÊNCIAS: python-docx, openpyxl, pywin32, pillow, lxml
"""

import multiprocessing
multiprocessing.freeze_support()

import os
import sys
import shutil
import zipfile
import tempfile
import threading
import traceback
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from PIL import Image, ImageDraw
import win32com.client
import pythoncom
from lxml import etree

# ─────────────────────────────────────────────
# PALETA
# ─────────────────────────────────────────────
BG       = "#0f1923"
BG2      = "#1a2535"
CAMPO    = "#243447"
ACENTO   = "#2e86de"
ACENTO2  = "#27ae60"
TEXTO    = "#e8edf3"
TEXTO2   = "#7a99b8"
BORDA    = "#2a3f58"
LOG_BG   = "#0a1118"
LOG_FG   = "#4ecb8d"

# ─────────────────────────────────────────────
# VALORES PADRÃO (copiar do app.py atual)
# ─────────────────────────────────────────────
DEFAULT = {
    # Assinatura
    "ass_ancora":    "AE72",
    "ass_offset_x":  10,
    "ass_offset_y":  -5,
    "ass_largura":   170,
    "ass_altura":    55,
    # Checkbox imagem — SIM (célula AM70)
    "chk_sim_ancora":   "AM70",
    "chk_sim_offset_x": 0,
    "chk_sim_offset_y": 0,
    "chk_sim_largura":  35,
    "chk_sim_altura":   11,
    # Checkbox imagem — NÃO (célula AP70)
    "chk_nao_ancora":   "AP70",
    "chk_nao_offset_x": 0,
    "chk_nao_offset_y": 0,
    "chk_nao_largura":  35,
    "chk_nao_altura":   11,
    "esgoto_sim":    True,
    # Geminadas
    "gem_cond":      "nao_se_aplica",
}

# Shapes de geminadas confirmados via drawing1.xml
SHAPES_GEMINADAS = {
    # Loteamentos (linha 64)
    "lot_sim":  "QOCI,13.L0C-32;L0C-34^",
    "lot_nao":  "QOCI,23.L0C-35;L0C-37^",
    "lot_nsa":  "QOCI,33.L0C-38;L0C-40^",
    # Condomínios (linha 65)
    "cond_sim": "QOCN,13.L0C-32;L0C-34^",
    "cond_nao": "QOCN,23.L0C-35;L0C-37^",
    "cond_nsa": "QOCN,33.L0C-38;L0C-40^",
}


# ─────────────────────────────────────────────
# HELPERS COM
# ─────────────────────────────────────────────

def _xls_para_xlsx_temp(path):
    if str(path).lower().endswith(".xlsx"):
        return str(path), False
    tmp = tempfile.mktemp(suffix=".xlsx")
    pythoncom.CoInitialize()
    xl = None; wb = None
    try:
        xl = win32com.client.Dispatch("Excel.Application")
        try: xl.Visible = False
        except: pass
        try: xl.DisplayAlerts = False
        except: pass
        wb = xl.Workbooks.Open(os.path.abspath(str(path)))
        wb.SaveAs(os.path.abspath(tmp), FileFormat=51)
        return tmp, True
    finally:
        if wb:
            try: wb.Close(SaveChanges=False)
            except: pass
        if xl:
            try: xl.Quit()
            except: pass
        pythoncom.CoUninitialize()


def _criar_xl():
    """Cria instância Excel limpa, matando zumbis se necessário."""
    import subprocess, time

    def _try():
        xl = win32com.client.Dispatch("Excel.Application")
        try: xl.Visible = False
        except: pass
        try: xl.DisplayAlerts = False
        except: pass
        try: xl.ScreenUpdating = False
        except: pass
        _ = xl.Workbooks.Count  # testa se está vivo
        return xl

    try:
        return _try()
    except Exception:
        pass
    try:
        subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"],
                       capture_output=True, creationflags=0x08000000)
    except Exception:
        pass
    time.sleep(1)
    return _try()


def _inserir_imagem_win32(ws, img_path, ancora, offset_x, offset_y, largura, altura):
    cell = ws.Range(ancora)
    left = cell.Left + offset_x
    top  = cell.Top  + offset_y
    ws.Shapes.AddPicture(
        os.path.abspath(img_path),
        False, True,
        left, top, largura, altura,
    )


# ─────────────────────────────────────────────
# GERAÇÃO DO PREVIEW
# ─────────────────────────────────────────────


# ─────────────────────────────────────────────
# CHECKBOX NATIVO (manipulação XML dos shapes)
# ─────────────────────────────────────────────

TEXTO_ITEM_ESGOTO = "sistema público de coleta de esgoto sanitário"


def _detectar_shapes_esgoto(xlsx_path, log):
    """Detecta nomes e posições dos shapes de esgoto no drawing XML."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path, data_only=False, read_only=True)
    sheet = "ElemConstrutivos" if "ElemConstrutivos" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    linha_esgoto = None
    col_sim = col_nao = None

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if TEXTO_ITEM_ESGOTO in cell.value.lower():
                    linha_esgoto = cell.row
                    break
        if linha_esgoto:
            break

    if not linha_esgoto:
        log("  ⚠ Texto de esgoto não encontrado no memorial")
        wb.close()
        return None, None, None, None

    for cell in ws[linha_esgoto]:
        if cell.value and isinstance(cell.value, str):
            v = cell.value.strip().upper()
            if v == "SIM" and col_sim is None:
                col_sim = cell.column
            elif v in ("NÃO", "NAO") and col_nao is None:
                col_nao = cell.column
    wb.close()

    ancora_sim = f"{get_column_letter(col_sim)}{linha_esgoto}" if col_sim else None
    ancora_nao = f"{get_column_letter(col_nao)}{linha_esgoto}" if col_nao else None

    shape_sim = shape_nao = None
    with zipfile.ZipFile(xlsx_path) as z:
        drawings = [f for f in z.namelist()
                    if f.startswith("xl/drawings/drawing") and f.endswith(".xml")]
        for drw in drawings:
            root = etree.fromstring(z.read(drw))
            ns = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
            nsm = {"xdr": ns}
            for anchor in root.findall("xdr:twoCellAnchor", nsm):
                sp = anchor.find("xdr:sp", nsm)
                if sp is None:
                    continue
                frm = anchor.find("xdr:from", nsm)
                if frm is None:
                    continue
                r = int(frm.find("xdr:row", nsm).text) + 1
                c = int(frm.find("xdr:col", nsm).text) + 1
                if r != linha_esgoto:
                    continue
                cnv = sp.find(f".//{{{ns}}}cNvPr")
                nome = cnv.get("name", "") if cnv is not None else ""
                if col_sim and c == col_sim:
                    shape_sim = nome
                elif col_nao and c == col_nao:
                    shape_nao = nome

    log(f"  ✓ Linha esgoto: {linha_esgoto} | SIM={ancora_sim}({shape_sim}) NÃO={ancora_nao}({shape_nao})")
    return shape_sim, shape_nao, ancora_sim, ancora_nao


def aplicar_checkbox_nativo(xlsx_path, esgoto_sim, log, gem_lot="nao_se_aplica", gem_cond="nao_se_aplica"):
    """Marca checkboxes manipulando diretamente o XML dos shapes."""
    import shutil as _sh

    shape_sim, shape_nao, ancora_sim, ancora_nao = _detectar_shapes_esgoto(xlsx_path, log)
    if not shape_sim and not shape_nao:
        log("  ✗ Nenhum shape encontrado — método nativo não disponível neste template")
        return False

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    _sh.copy2(xlsx_path, tmp_path)

    shapes_ok = 0
    try:
        with zipfile.ZipFile(tmp_path, "r") as zi,              zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zo:
            for item in zi.infolist():
                data = zi.read(item.filename)
                if item.filename.startswith("xl/drawings/drawing") and item.filename.endswith(".xml"):
                    try:
                        root = etree.fromstring(data)
                        nsm = {
                            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                            "a":   "http://schemas.openxmlformats.org/drawingml/2006/main",
                        }
                        for sp in root.iter("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}sp"):
                            cnv = sp.find(".//xdr:nvSpPr/xdr:cNvPr", nsm)
                            if cnv is None:
                                continue
                            nome = cnv.get("name", "")
                            cor = None
                            if shape_sim and nome == shape_sim:
                                cor = "000000" if esgoto_sim else "FFFFFF"
                            elif shape_nao and nome == shape_nao:
                                cor = "000000" if not esgoto_sim else "FFFFFF"
                            # Geminadas loteamentos
                            elif nome == SHAPES_GEMINADAS["lot_sim"]:
                                cor = "000000" if gem_lot == "sim" else "FFFFFF"
                            elif nome == SHAPES_GEMINADAS["lot_nao"]:
                                cor = "000000" if gem_lot == "nao" else "FFFFFF"
                            elif nome == SHAPES_GEMINADAS["lot_nsa"]:
                                cor = "000000" if gem_lot == "nao_se_aplica" else "FFFFFF"
                            # Geminadas condomínios
                            elif nome == SHAPES_GEMINADAS["cond_sim"]:
                                cor = "000000" if gem_cond == "sim" else "FFFFFF"
                            elif nome == SHAPES_GEMINADAS["cond_nao"]:
                                cor = "000000" if gem_cond == "nao" else "FFFFFF"
                            elif nome == SHAPES_GEMINADAS["cond_nsa"]:
                                cor = "000000" if gem_cond == "nao_se_aplica" else "FFFFFF"
                            if cor is None:
                                continue
                            shapes_ok += 1
                            sp_pr = sp.find(".//xdr:spPr", nsm)
                            if sp_pr is None:
                                continue
                            for ft in ["a:solidFill", "a:noFill", "a:gradFill"]:
                                ex = sp_pr.find(ft, nsm)
                                if ex is not None:
                                    sp_pr.remove(ex)
                            sf = etree.SubElement(
                                sp_pr,
                                "{http://schemas.openxmlformats.org/drawingml/2006/main}solidFill"
                            )
                            sc = etree.SubElement(
                                sf,
                                "{http://schemas.openxmlformats.org/drawingml/2006/main}srgbClr"
                            )
                            sc.set("val", cor)
                            xfrm = sp_pr.find("a:xfrm", nsm)
                            sp_pr.remove(sf)
                            if xfrm is not None:
                                xfrm.addnext(sf)
                            else:
                                sp_pr.insert(0, sf)
                        data = etree.tostring(root, xml_declaration=True,
                                              encoding="UTF-8", standalone=True)
                    except Exception as e:
                        log(f"  ⚠ Erro XML: {e}")
                zo.writestr(item, data)
        os.unlink(tmp_path)
        if shapes_ok == 0:
            log("  ⚠ Shapes encontrados no XML mas nenhuma cor foi alterada")
            return False
        log(f"  ✓ {shapes_ok} shape(s) modificado(s) via método NATIVO")
        return True
    except Exception as e:
        log(f"  ✗ Falha no método nativo: {e}")
        try: os.unlink(tmp_path)
        except: pass
        return False

def _fechar_excel(xl, wb):
    """Fecha wb e xl de forma segura, mata EXCEL.EXE se necessário."""
    if wb is not None:
        try: wb.Close(SaveChanges=False)
        except Exception: pass
    if xl is not None:
        try: xl.Quit()
        except Exception: pass
        try:
            import subprocess
            subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"],
                           capture_output=True, creationflags=0x08000000)
        except Exception: pass


def gerar_preview(memorial_path, ass_img_path, cfg, saida_path, log, modo_checkbox="imagem"):
    """
    Abre o memorial, insere assinatura + checkbox de preview e exporta PDF.
    Roda em thread separada.
    """
    pythoncom.CoInitialize()
    xl = None; wb = None
    xlsx_tmp = None; xlsx_criou = False

    try:
        # 1. Garantir .xlsx — nunca copiar sobre si mesmo
        log("• Preparando arquivo...")
        xlsx_tmp, xlsx_criou = _xls_para_xlsx_temp(memorial_path)
        src_real  = os.path.realpath(xlsx_tmp)
        dest_real = os.path.realpath(saida_path)
        if src_real == dest_real:
            # Copiar para temp intermediário primeiro
            import tempfile as _tmp
            inter = _tmp.mktemp(suffix=".xlsx")
            shutil.copy2(xlsx_tmp, inter)
            if xlsx_criou:
                try: os.unlink(xlsx_tmp)
                except: pass
            shutil.move(inter, saida_path)
        else:
            shutil.copy2(xlsx_tmp, saida_path)
            if xlsx_criou:
                try: os.unlink(xlsx_tmp)
                except: pass

        # 2. Abrir no Excel
        log("• Abrindo no Excel...")
        xl = _criar_xl()
        wb = xl.Workbooks.Open(os.path.abspath(saida_path))
        try:
            ws = wb.Worksheets("ElemConstrutivos")
        except Exception:
            ws = wb.Worksheets(1)

        # 3. Inserir assinatura
        if ass_img_path and os.path.exists(ass_img_path):
            log(f"• Inserindo assinatura em {cfg['ass_ancora']} "
                f"({cfg['ass_largura']}×{cfg['ass_altura']}pt)...")
            _inserir_imagem_win32(
                ws, ass_img_path,
                cfg["ass_ancora"],
                cfg["ass_offset_x"], cfg["ass_offset_y"],
                cfg["ass_largura"],  cfg["ass_altura"],
            )
        else:
            log("  ⚠ Assinatura não selecionada — pulando")

        # 4. Checkbox — nativo ou imagem
        esgoto_sim = cfg["esgoto_sim"]
        if modo_checkbox == "nativo":
            # Nativo: salvar SEM shapes extras → fechar tudo → modificar XML → reabrir
            log("• Salvando antes de aplicar checkboxes nativos...")
            wb.Save()
            _fechar_excel(xl, wb)
            xl = None; wb = None
            pythoncom.CoUninitialize()
            import time; time.sleep(1)
            pythoncom.CoInitialize()

            log("• Aplicando checkboxes via método NATIVO (XML)...")
            gem_lot  = cfg.get("gem_lot",  "nao_se_aplica")
            gem_cond = cfg.get("gem_cond", "nao_se_aplica")
            ok = aplicar_checkbox_nativo(saida_path, esgoto_sim, log,
                                         gem_lot=gem_lot, gem_cond=gem_cond)
            if not ok:
                log("  ⚠ Nativo falhou — shapes não encontrados neste template")

            # Reabrir para exportar PDF
            xl = _criar_xl()
            wb = xl.Workbooks.Open(os.path.abspath(saida_path))
            try:
                ws = wb.Worksheets("ElemConstrutivos")
            except Exception:
                ws = wb.Worksheets(1)
        else:
            chk_img = _obter_img_checkbox(esgoto_sim)
            chk_marcado = _obter_img_quadrado_preto()
            if esgoto_sim:
                log(f"• Checkbox SIM em {cfg['chk_sim_ancora']} "
                    f"({cfg['chk_sim_largura']}×{cfg['chk_sim_altura']}pt)...")
                _inserir_imagem_win32(ws, chk_marcado,
                    cfg["chk_sim_ancora"],
                    cfg["chk_sim_offset_x"], cfg["chk_sim_offset_y"],
                    cfg["chk_sim_largura"],  cfg["chk_sim_altura"])
            else:
                log(f"• Checkbox NÃO em {cfg['chk_nao_ancora']} "
                    f"({cfg['chk_nao_largura']}×{cfg['chk_nao_altura']}pt)...")
                _inserir_imagem_win32(ws, chk_marcado,
                    cfg["chk_nao_ancora"],
                    cfg["chk_nao_offset_x"], cfg["chk_nao_offset_y"],
                    cfg["chk_nao_largura"],  cfg["chk_nao_altura"])

        # 5. Salvar e exportar PDF
        wb.Save()
        pdf_path = saida_path.replace(".xlsx", ".pdf")
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = 1
        ws.ExportAsFixedFormat(0, os.path.abspath(pdf_path), 0, True, False)

        log(f"✅ PDF gerado: {pdf_path}")
        log("─" * 50)
        log("VALORES PARA COPIAR PARA O app.py:")
        log(f'  ASSINATURA_EXCEL_ANCORA     = "{cfg["ass_ancora"]}"')
        log(f'  ASSINATURA_EXCEL_OFFSET_X_PT = {cfg["ass_offset_x"]}')
        log(f'  ASSINATURA_EXCEL_OFFSET_Y_PT = {cfg["ass_offset_y"]}')
        log(f'  ASSINATURA_EXCEL_LARGURA_PT  = {cfg["ass_largura"]}')
        log(f'  ASSINATURA_EXCEL_ALTURA_PT   = {cfg["ass_altura"]}')
        log(f'  # Checkbox SIM')
        log(f'  CHECKBOX_SIM_ANCORA      = "{cfg["chk_sim_ancora"]}"')
        log(f'  CHECKBOX_SIM_OFFSET_X_PT = {cfg["chk_sim_offset_x"]}')
        log(f'  CHECKBOX_SIM_OFFSET_Y_PT = {cfg["chk_sim_offset_y"]}')
        log(f'  CHECKBOX_SIM_LARGURA_PT  = {cfg["chk_sim_largura"]}')
        log(f'  CHECKBOX_SIM_ALTURA_PT   = {cfg["chk_sim_altura"]}')
        log(f'  # Checkbox NÃO')
        log(f'  CHECKBOX_NAO_ANCORA      = "{cfg["chk_nao_ancora"]}"')
        log(f'  CHECKBOX_NAO_OFFSET_X_PT = {cfg["chk_nao_offset_x"]}')
        log(f'  CHECKBOX_NAO_OFFSET_Y_PT = {cfg["chk_nao_offset_y"]}')
        log(f'  CHECKBOX_NAO_LARGURA_PT  = {cfg["chk_nao_largura"]}')
        log(f'  CHECKBOX_NAO_ALTURA_PT   = {cfg["chk_nao_altura"]}')

        # Abrir PDF automaticamente
        try:
            os.startfile(pdf_path)
        except Exception:
            pass

    except Exception as e:
        log(f"✗ ERRO: {e}")
        log(traceback.format_exc())
    finally:
        _fechar_excel(xl, wb)
        pythoncom.CoUninitialize()


def _obter_img_quadrado_preto():
    """Quadrado preto sólido — replica o shape marcado (■)."""
    tmp = tempfile.mktemp(suffix=".png")
    Image.new("RGBA", (20, 20), (0, 0, 0, 255)).save(tmp)
    return tmp


def _obter_img_checkbox(esgoto_sim):

    """
    Retorna caminho de imagem de checkbox.
    Tenta encontrar no diretório do script; se não, cria placeholder.
    """
    base = Path(__file__).parent
    candidatos_sim = [
        base / "assets" / "CHECKBOX_COM_ESGOTO.png",
        base / "assets" / "CHECKBOX_COM_ESGOTO.jpeg",
        base / "assets" / "CHECKBOX_COM_ESGOTO.png.jpeg",
    ]
    candidatos_nao = [
        base / "assets" / "CHECKBOX_SEM_ESGOTO.png",
        base / "assets" / "CHECKBOX_SEM_ESGOTO.jpeg",
        base / "assets" / "CHECKBOX_SEM_ESGOTO.png.jpeg",
    ]
    lista = candidatos_sim if esgoto_sim else candidatos_nao
    for c in lista:
        if c.exists():
            return str(c)

    # Placeholder: dois quadrados — o marcado fica preto (igual ao shape original)
    # Dimensões proporcionais a ~14pt altura x 85pt largura
    W, H = 170, 28
    SZ = H - 4  # tamanho do quadrado
    tmp = tempfile.mktemp(suffix=".png")
    img = Image.new("RGBA", (W, H), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    # Quadrado SIM (esquerda)
    sim_x = 2
    if esgoto_sim:
        draw.rectangle([sim_x, 2, sim_x + SZ, 2 + SZ], fill=(0, 0, 0, 255))   # ■ preto
    else:
        draw.rectangle([sim_x, 2, sim_x + SZ, 2 + SZ], outline=(0, 0, 0, 255), width=2)  # □

    # Quadrado NÃO (direita)
    nao_x = W // 2 + 4
    if not esgoto_sim:
        draw.rectangle([nao_x, 2, nao_x + SZ, 2 + SZ], fill=(0, 0, 0, 255))   # ■ preto
    else:
        draw.rectangle([nao_x, 2, nao_x + SZ, 2 + SZ], outline=(0, 0, 0, 255), width=2)  # □

    img.save(tmp)
    return tmp


# ─────────────────────────────────────────────
# INTERFACE
# ─────────────────────────────────────────────

class Calibrador(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Calibrador do Memorial — Morais Engenharia")
        self.geometry("860x780")
        self.configure(bg=BG)
        self.resizable(True, True)
        self._criar_ui()

    # ── UI ──────────────────────────────────

    def _criar_ui(self):
        # ── Cabeçalho fixo ──
        hdr = tk.Frame(self, bg=BG, pady=14)
        hdr.pack(fill="x", padx=24)
        tk.Label(hdr, text="CALIBRADOR DO MEMORIAL",
                 font=("Segoe UI", 16, "bold"), fg=TEXTO, bg=BG).pack(anchor="w")
        tk.Label(hdr, text="Ajuste assinatura e checkbox • copie os valores para o app.py",
                 font=("Segoe UI", 9), fg=TEXTO2, bg=BG).pack(anchor="w")
        tk.Frame(self, bg=BORDA, height=1).pack(fill="x")

        # ── Botão GERAR PREVIEW fixo no rodapé — sempre visível ──
        rodape = tk.Frame(self, bg=BG, pady=10)
        rodape.pack(side="bottom", fill="x", padx=24)
        tk.Frame(rodape, bg=BORDA, height=1).pack(fill="x", pady=(0, 8))
        tk.Label(rodape,
                 text="💡 Teste os dois métodos de checkbox e compare o resultado no PDF",
                 bg=BG, fg=ACENTO2, font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 4))
        self.btn = tk.Button(rodape, text="⚡  GERAR PREVIEW (abre PDF automaticamente)",
                             command=self._iniciar,
                             bg=ACENTO, fg=TEXTO, relief="flat",
                             font=("Segoe UI", 12, "bold"), pady=10)
        self.btn.pack(fill="x")
        self.var_status = tk.StringVar(value="Aguardando...")
        tk.Label(rodape, textvariable=self.var_status,
                 bg=BG, fg=TEXTO2, font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 0))

        # ── Corpo principal ──
        tk.Frame(self, bg=BORDA, height=1).pack(fill="x")
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=0, pady=0)

        # Coluna esquerda COM scroll
        frame_esq_outer = tk.Frame(body, bg=BG, width=320)
        frame_esq_outer.pack(side="left", fill="y", padx=(24, 0))
        frame_esq_outer.pack_propagate(False)

        canvas_esq = tk.Canvas(frame_esq_outer, bg=BG, highlightthickness=0)
        sb_esq = tk.Scrollbar(frame_esq_outer, orient="vertical", command=canvas_esq.yview)
        canvas_esq.configure(yscrollcommand=sb_esq.set)
        sb_esq.pack(side="right", fill="y")
        canvas_esq.pack(side="left", fill="both", expand=True)

        col_esq = tk.Frame(canvas_esq, bg=BG)
        win_id = canvas_esq.create_window((0, 0), window=col_esq, anchor="nw")

        def _on_configure(event):
            canvas_esq.configure(scrollregion=canvas_esq.bbox("all"))
            canvas_esq.itemconfig(win_id, width=canvas_esq.winfo_width())
        col_esq.bind("<Configure>", _on_configure)

        # Scroll com mouse na coluna esquerda
        def _on_mousewheel(event):
            canvas_esq.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas_esq.bind("<MouseWheel>", _on_mousewheel)
        col_esq.bind("<MouseWheel>", _on_mousewheel)

        # Coluna direita fixa
        col_dir = tk.Frame(body, bg=BG)
        col_dir.pack(side="right", fill="both", expand=True, padx=(10, 24))

        # ── Coluna esquerda ──
        self._titulo(col_esq, "ARQUIVOS")
        self.var_memorial = tk.StringVar()
        self._arquivo(col_esq, self.var_memorial, "Memorial Excel (.xls/.xlsx)",
                      [("Excel", "*.xls *.xlsx")])
        self.var_assinatura = tk.StringVar()
        self._arquivo(col_esq, self.var_assinatura, "Imagem da assinatura (.png/.jpg)",
                      [("Imagem", "*.png *.jpg *.jpeg")])

        self._titulo(col_esq, "ASSINATURA NO MEMORIAL")
        self._linha_campo(col_esq, "Célula âncora", "ass_ancora",  DEFAULT["ass_ancora"])
        self._linha_spin(col_esq,  "Offset X (pt)",  "ass_off_x",  DEFAULT["ass_offset_x"], -200, 200)
        self._linha_spin(col_esq,  "Offset Y (pt)",  "ass_off_y",  DEFAULT["ass_offset_y"], -200, 200)
        self._linha_spin(col_esq,  "Largura (pt)",   "ass_larg",   DEFAULT["ass_largura"],   10, 400)
        self._linha_spin(col_esq,  "Altura (pt)",    "ass_alt",    DEFAULT["ass_altura"],    10, 200)

        tk.Label(col_esq, text="1 cm ≈ 28 pt  |  1 polegada = 72 pt",
                 bg=BG, fg=TEXTO2, font=("Segoe UI", 8)).pack(anchor="w", pady=(2, 8))

        self._titulo(col_esq, "CHECKBOX DE ESGOTO")
        self.var_esgoto = tk.BooleanVar(value=DEFAULT["esgoto_sim"])
        tk.Checkbutton(col_esq, text="Sistema público de esgoto (SIM)",
                       variable=self.var_esgoto,
                       bg=BG, fg=TEXTO, selectcolor=CAMPO,
                       activebackground=BG, activeforeground=TEXTO,
                       font=("Segoe UI", 9)).pack(anchor="w", pady=4)

        self._titulo(col_esq, "CASAS GEMINADAS (método nativo)")
        _og = ["Não se aplica", "Sim", "Não"]
        tk.Label(col_esq, text="Loteamentos (linha 64)",
                 bg=BG, fg=TEXTO2, font=("Segoe UI", 8)).pack(anchor="w")
        self.var_gem_lot = tk.StringVar(value="Não se aplica")
        ttk.Combobox(col_esq, textvariable=self.var_gem_lot, values=_og,
                     state="readonly", font=("Segoe UI", 9)).pack(fill="x", pady=(0,6))
        tk.Label(col_esq, text="Condomínios (linha 65)",
                 bg=BG, fg=TEXTO2, font=("Segoe UI", 8)).pack(anchor="w")
        self.var_gem_cond = tk.StringVar(value="Não se aplica")
        ttk.Combobox(col_esq, textvariable=self.var_gem_cond, values=_og,
                     state="readonly", font=("Segoe UI", 9)).pack(fill="x", pady=(0,6))

        self._titulo(col_esq, "MÉTODO DO CHECKBOX")
        self.var_modo_checkbox = tk.StringVar(value="imagem")
        fr_mc = tk.Frame(col_esq, bg=BG)
        fr_mc.pack(anchor="w", pady=4)
        for txt, val in [("Imagem (sobrepõe PNG)", "imagem"),
                          ("Nativo (manipula shapes XML)", "nativo")]:
            tk.Radiobutton(
                fr_mc, text=txt, variable=self.var_modo_checkbox, value=val,
                bg=BG, fg=TEXTO, selectcolor=CAMPO,
                activebackground=BG, activeforeground=TEXTO,
                font=("Segoe UI", 9),
            ).pack(anchor="w")


        # ── Coluna direita: log ──
        self._titulo(col_dir, "LOG")
        frame_log = tk.Frame(col_dir, bg=LOG_BG)
        frame_log.pack(fill="both", expand=True, pady=4)
        sb = tk.Scrollbar(frame_log)
        sb.pack(side="right", fill="y")
        self.txt_log = tk.Text(frame_log, bg=LOG_BG, fg=LOG_FG,
                               font=("Consolas", 9), relief="flat",
                               yscrollcommand=sb.set, wrap="word")
        self.txt_log.pack(fill="both", expand=True)
        sb.config(command=self.txt_log.yview)

        self._titulo(col_dir, "VALORES PARA COPIAR")
        self.txt_copy = tk.Text(col_dir, height=12, bg=CAMPO, fg=ACENTO2,
                                font=("Consolas", 9), relief="flat")
        self.txt_copy.pack(fill="x", pady=4)
        self._atualizar_copy()

        tk.Button(col_dir, text="📋  COPIAR VALORES",
                  command=self._copiar_valores,
                  bg=BORDA, fg=TEXTO, relief="flat",
                  font=("Segoe UI", 9, "bold"), pady=6).pack(fill="x", pady=(0, 8))



        # Vincular atualização automática dos valores
        for v in self._vars.values():
            v.trace_add("write", lambda *_: self.after(100, self._atualizar_copy))
        self.var_esgoto.trace_add("write", lambda *_: self.after(100, self._atualizar_copy))

    # ── Helpers de UI ───────────────────────

    def _titulo(self, parent, txt):
        tk.Label(parent, text=txt, font=("Segoe UI", 10, "bold"),
                 fg=TEXTO, bg=BG).pack(anchor="w", pady=(12, 4))

    def _arquivo(self, parent, var, hint, ftypes):
        tk.Label(parent, text=hint, fg=TEXTO2, bg=BG,
                 font=("Segoe UI", 8)).pack(anchor="w")
        fr = tk.Frame(parent, bg=BG)
        fr.pack(fill="x", pady=(0, 4))
        tk.Entry(fr, textvariable=var, bg=CAMPO, fg=TEXTO,
                 insertbackground=TEXTO, relief="flat",
                 font=("Segoe UI", 9)).pack(side="left", fill="x", expand=True)
        tk.Button(fr, text="📁",
                  command=lambda: self._sel(var, ftypes),
                  bg=ACENTO, fg=TEXTO, relief="flat",
                  font=("Segoe UI", 9)).pack(side="right", padx=(4, 0))

    def _sel(self, var, ftypes):
        p = filedialog.askopenfilename(filetypes=ftypes)
        if p:
            var.set(p)

    _vars: dict = {}

    def _linha_campo(self, parent, label, key, default):
        if not hasattr(self, "_vars"):
            self._vars = {}
        fr = tk.Frame(parent, bg=BG)
        fr.pack(fill="x", pady=2)
        tk.Label(fr, text=label, width=18, anchor="w",
                 fg=TEXTO2, bg=BG, font=("Segoe UI", 9)).pack(side="left")
        var = tk.StringVar(value=str(default))
        self._vars[key] = var
        tk.Entry(fr, textvariable=var, width=10, bg=CAMPO, fg=TEXTO,
                 insertbackground=TEXTO, relief="flat",
                 font=("Segoe UI", 10, "bold")).pack(side="left")

    def _linha_spin(self, parent, label, key, default, mn, mx):
        if not hasattr(self, "_vars"):
            self._vars = {}
        fr = tk.Frame(parent, bg=BG)
        fr.pack(fill="x", pady=2)
        tk.Label(fr, text=label, width=18, anchor="w",
                 fg=TEXTO2, bg=BG, font=("Segoe UI", 9)).pack(side="left")
        var = tk.StringVar(value=str(default))
        self._vars[key] = var
        tk.Spinbox(fr, textvariable=var, from_=mn, to=mx, width=8,
                   bg=CAMPO, fg=TEXTO, insertbackground=TEXTO,
                   buttonbackground=BORDA, relief="flat",
                   font=("Segoe UI", 10, "bold")).pack(side="left")
        # botões ±5
        tk.Button(fr, text="-5",
                  command=lambda k=key, v=var: self._nudge(v, -5),
                  bg=BORDA, fg=TEXTO2, relief="flat",
                  font=("Segoe UI", 8), padx=4).pack(side="left", padx=(6, 1))
        tk.Button(fr, text="+5",
                  command=lambda k=key, v=var: self._nudge(v, +5),
                  bg=BORDA, fg=TEXTO2, relief="flat",
                  font=("Segoe UI", 8), padx=4).pack(side="left", padx=1)

    def _nudge(self, var, delta):
        try:
            var.set(str(int(var.get()) + delta))
        except Exception:
            pass

    def _cfg(self):
        v = self._vars
        def i(k): return int(v[k].get())
        def s(k): return v[k].get().strip()
        return {
            "ass_ancora":       s("ass_ancora"),
            "ass_offset_x":     i("ass_off_x"),
            "ass_offset_y":     i("ass_off_y"),
            "ass_largura":      i("ass_larg"),
            "ass_altura":       i("ass_alt"),
            "chk_sim_ancora":   s("chk_sim_ancora"),
            "chk_sim_offset_x": i("chk_sim_off_x"),
            "chk_sim_offset_y": i("chk_sim_off_y"),
            "chk_sim_largura":  i("chk_sim_larg"),
            "chk_sim_altura":   i("chk_sim_alt"),
            "chk_nao_ancora":   s("chk_nao_ancora"),
            "chk_nao_offset_x": i("chk_nao_off_x"),
            "chk_nao_offset_y": i("chk_nao_off_y"),
            "chk_nao_largura":  i("chk_nao_larg"),
            "chk_nao_altura":   i("chk_nao_alt"),
            "esgoto_sim":       self.var_esgoto.get(),
            "gem_cond":         {"Não se aplica": "nao_se_aplica",
                                 "Sim": "sim", "Não": "nao"}.get(
                                 self.var_gem_cond.get(), "nao_se_aplica"),
        }

    def _atualizar_copy(self):
        try:
            cfg = self._cfg()
        except Exception:
            return
        txt = (
            f'ASSINATURA_EXCEL_ANCORA      = "{cfg["ass_ancora"]}"\n'
            f'ASSINATURA_EXCEL_OFFSET_X_PT = {cfg["ass_offset_x"]}\n'
            f'ASSINATURA_EXCEL_OFFSET_Y_PT = {cfg["ass_offset_y"]}\n'
            f'ASSINATURA_EXCEL_LARGURA_PT  = {cfg["ass_largura"]}\n'
            f'ASSINATURA_EXCEL_ALTURA_PT   = {cfg["ass_altura"]}\n'
            f'# Checkbox SIM\n'
            f'CHECKBOX_SIM_ANCORA      = "{cfg["chk_sim_ancora"]}"\n'
            f'CHECKBOX_SIM_OFFSET_X_PT = {cfg["chk_sim_offset_x"]}\n'
            f'CHECKBOX_SIM_OFFSET_Y_PT = {cfg["chk_sim_offset_y"]}\n'
            f'CHECKBOX_SIM_LARGURA_PT  = {cfg["chk_sim_largura"]}\n'
            f'CHECKBOX_SIM_ALTURA_PT   = {cfg["chk_sim_altura"]}\n'
            f'# Checkbox NÃO\n'
            f'CHECKBOX_NAO_ANCORA      = "{cfg["chk_nao_ancora"]}"\n'
            f'CHECKBOX_NAO_OFFSET_X_PT = {cfg["chk_nao_offset_x"]}\n'
            f'CHECKBOX_NAO_OFFSET_Y_PT = {cfg["chk_nao_offset_y"]}\n'
            f'CHECKBOX_NAO_LARGURA_PT  = {cfg["chk_nao_largura"]}\n'
            f'CHECKBOX_NAO_ALTURA_PT   = {cfg["chk_nao_altura"]}\n'
        )
        self.txt_copy.delete("1.0", "end")
        self.txt_copy.insert("1.0", txt)

    def _copiar_valores(self):
        self.clipboard_clear()
        self.clipboard_append(self.txt_copy.get("1.0", "end").strip())
        self.var_status.set("✓ Valores copiados para a área de transferência!")

    # ── Ação principal ──────────────────────

    def _iniciar(self):
        memorial = self.var_memorial.get().strip()
        if not memorial or not os.path.exists(memorial):
            messagebox.showerror("Erro", "Selecione um arquivo Memorial válido.")
            return
        if "PREVIEW_MEMORIAL_CALIBRADOR" in os.path.basename(memorial):
            messagebox.showerror(
                "Arquivo inválido",
                "Você selecionou o arquivo de PREVIEW gerado pelo calibrador.\n\n"
                "Selecione o memorial ORIGINAL (.xls ou .xlsx)."
            )
            return
        try:
            cfg = self._cfg()
        except ValueError as e:
            messagebox.showerror("Erro", f"Valor inválido: {e}")
            return

        ass_img = self.var_assinatura.get().strip() or None

        saida = str(Path(memorial).parent / "PREVIEW_MEMORIAL_CALIBRADOR.xlsx")

        self.btn.configure(state="disabled", text="⏳  Gerando preview...")
        self.txt_log.delete("1.0", "end")
        self.var_status.set("Processando...")

        modo_checkbox = self.var_modo_checkbox.get()
        threading.Thread(
            target=self._worker,
            args=(memorial, ass_img, cfg, saida, modo_checkbox),
            daemon=True,
        ).start()

    def _worker(self, memorial, ass_img, cfg, saida, modo_checkbox="imagem"):
        def log(msg):
            self.after(0, self._log_insert, msg)

        try:
            gerar_preview(memorial, ass_img, cfg, saida, log, modo_checkbox=modo_checkbox)
            self.after(0, self.var_status.set, "✅ Preview gerado — verifique o PDF!")
        except Exception as e:
            self.after(0, self.var_status.set, f"✗ Erro: {e}")
            self.after(0, self._log_insert, f"✗ ERRO FATAL: {e}")
            # Garantir kill do Excel mesmo em erro não capturado
            try:
                import subprocess
                subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"],
                               capture_output=True, creationflags=0x08000000)
            except Exception: pass
        finally:
            self.after(0, self.btn.configure, {
                "state": "normal",
                "text": "⚡  GERAR PREVIEW (abre PDF automaticamente)"
            })

    def _log_insert(self, msg):
        self.txt_log.insert("end", msg + "\n")
        self.txt_log.see("end")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    Calibrador().mainloop()
